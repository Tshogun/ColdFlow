import sys
import os
import openpyxl
from datetime import datetime, timedelta

# Add logger directory to sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../logger')))
from logger import Logger
log = Logger()

# Add llm directory to sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../llm')))
from generate_content import CoverLetterGenerator  # type: ignore
generator = CoverLetterGenerator()

# Add email directory to sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../email')))
from send_email import EmailSender, getenv  # type: ignore
sender_email, password = getenv()
es = EmailSender(sender_email, password)
attachment_path = "src/resume/document.pdf"

class ColdFlow:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.sheet = None
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file_path)
            self.sheet = self.workbook.active
            log.info(f"Loaded workbook: {self.excel_file_path}, sheet: {self.sheet.title}")
        except FileNotFoundError:
            log.error(f"Error: Excel file not found at {self.excel_file_path}")
        except Exception as e:
            log.error(f"Error loading workbook: {e}", exc_info=True)

    def process_excel_sheet(self):
        """
        Iterates through each row (skipping header, excluding first column),
        processes emails, and updates the Excel sheet.
        """
        if not self.sheet:
            log.error("Sheet not loaded. Cannot process.")
            return

        header_skipped = False
        row_index_excel = 1  # Start with the first row in Excel

        for row in self.sheet.iter_rows(values_only=True):
            if not header_skipped:
                log.debug("Skipping header row.")
                header_skipped = True
                row_index_excel += 1
                continue

            row_data = list(row[1:])  # Skip the first column
            if len(row_data) < 10:
                log.warning(f"Skipping row {row_index_excel} due to insufficient data: {row_data}")
                row_index_excel += 1
                continue

            company_name = row_data[0]
            email_recipient = row_data[1]
            email_id = row_data[2]
            send_count_str = row_data[3]
            frequency_str = row_data[4]
            last_email_date_raw = row_data[6]
            prompt = row_data[7]
            email_content = row_data[8]
            status = row_data[9]

            log.info(f"Processing row: {email_recipient} at {company_name} (Excel Row: {row_index_excel})")
            try:
                # Check if the values are not None before attempting conversion
                if send_count_str is not None and frequency_str is not None:
                    send_count = int(send_count_str)
                    frequency = int(frequency_str)
                else:
                    log.error(f"Send_Count or Frequency is None for {email_recipient}. Skipping row.")
                    row_index_excel += 1
                    continue
            except ValueError:
                log.error(f"Could not convert Send_Count '{send_count_str}' or Frequency '{frequency_str}' to integer for {email_recipient}. Skipping row.")
                row_index_excel += 1
                continue

            if send_count > 1:
                log.debug(f"Send_Count ({send_count}) is greater than 1. Proceeding with email check.")

                send_email_flag = False

                if last_email_date_raw is None:
                    log.debug("Last_Email_Date is None. Proceeding to send email.")
                    send_email_flag = True
                else:
                    today = datetime.now().date()
                    last_email_date = None
                    if isinstance(last_email_date_raw, datetime):
                        last_email_date = last_email_date_raw.date()
                    else:
                        try:
                            last_email_date = datetime.strptime(str(last_email_date_raw).split()[0], '%Y-%m-%d').date()
                        except (ValueError, AttributeError):
                            log.warning(f"Invalid Last_Email_Date format: {last_email_date_raw}. Proceeding to send email.")
                            send_email_flag = True  # Treat as no last email date

                    if last_email_date:
                        time_difference = today - last_email_date
                        if time_difference >= timedelta(days=frequency):
                            log.debug(f"Time difference ({time_difference}) matches/exceeds frequency ({frequency}). Proceeding to send email.")
                            send_email_flag = True
                        else:
                            log.debug(f"Time difference ({time_difference}) does not match frequency ({frequency}). Skipping email for this row.")

                if send_email_flag:
                    if email_content:
                        es.send_email(email_id, "Internship Application", email_content, attachment_path)
                        log.debug(f"Successfully sent email to {email_recipient} at {company_name}")
                        self.sheet.cell(row=row_index_excel + 1, column=9, value=email_content) # Update Email Content
                    else:
                        email_content = generator.generate_cover_letter_first_time(prompt)
                        es.send_email(email_id, "Internship Application", email_content, attachment_path)
                        log.debug(f"Generated and sent email to {email_recipient} at {company_name}")
                        self.sheet.cell(row=row_index_excel, column=10, value=email_content) # Update Email Content

                    self.sheet.cell(row=row_index_excel, column=5, value=send_count - 1)
                    log.debug(f"Updated Send_Count to {send_count - 1} for row {row_index_excel}")
                    self.sheet.cell(row=row_index_excel, column=8, value=datetime.now().strftime('%Y-%m-%d'))
                    log.debug(f"Updated Last_Email_Date to {datetime.now().strftime('%Y-%m-%d')} for row {row_index_excel}")
                    self.sheet.cell(row=row_index_excel, column=11, value=f"Email Sent.")
                    log.debug(f"Updated Status to 'Email Sent' for row {row_index_excel}")

            row_index_excel += 1

    def save_workbook(self):
        if self.workbook:
            try:
                self.workbook.save(self.excel_file_path)
                log.info(f"Successfully saved changes to: {self.excel_file_path}")
                return True
            except Exception as e:
                log.error(f"Error saving workbook: {e}", exc_info=True)
                return False
        else:
            log.error("Workbook not loaded. Cannot save.")
            return False

if __name__ == "__main__":
    excel_file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', 'data', 'cold_email_data.xlsx')
    coldflow = ColdFlow(excel_file_path)
    coldflow.process_excel_sheet()
    coldflow.save_workbook()